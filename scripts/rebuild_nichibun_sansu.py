#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""日本文教出版 算数を「小単元＝単元」へ再構築する。

原本: textbooks/sansu/download/r6/r6_sansu_nenkei_{g}.xlsx
構造: 1シート = 1大単元
  r1  大単元名 「2　体積〔…〕 （p.17～31）」
  r2  時期と配当 「4月中旬～4月下旬［11時間］」
  以降 A列に小単元見出し 「1　直方体と立方体の体積（p.18～22）　3時間」
      「学習をたしかに（p.30～31）　1時間」なども同形式

方針: A列の「… N時間」形式の行を小単元として拾い、単元名は 大単元「小単元」 の入れ子に。
      小単元の合計が大単元の配当と一致しない場合は分割せず大単元のまま（推測しない）。
      総時数は既存テンプレート（＝大単元の配当）を正とする。

使い方: python3 scripts/rebuild_nichibun_sansu.py [--apply]
"""
import argparse
import json
import os
import re

import openpyxl

Z = "０１２３４５６７８９"
ROOT = os.path.join(os.path.dirname(__file__), "..")
TPL = os.path.join(ROOT, "client/public/templates/teaching")
SRC = os.path.join(ROOT, "sources/teaching-plans/小学校/算数/日本文教出版")

HDR_HOURS = re.compile(r"[［\[]\s*(\d+)\s*時間")          # 大単元の配当 ［11時間］
SUB = re.compile(r"^(.+?)\s*(?:（[^）]*）)?\s*(\d+)\s*時間\s*$")  # 小単元 … 3時間


def n(s):
    return (str(s) if s is not None else "").translate(str.maketrans(Z, "0123456789"))


def clean(name):
    name = re.sub(r"\s*（[^）]*）\s*$", "", name).strip()      # 末尾のページ表記
    name = re.sub(r"〔[^〕]*〕", "", name).strip()             # 副題〔…〕
    name = re.sub(r"^[●○◆]\s*", "", name).strip()
    return re.sub(r"\s+", " ", name)


def parse_sheet(ws):
    """(大単元名, 配当時数, [(小単元名, 時数)]) を返す。"""
    big = clean(n(ws.cell(1, 1).value))
    hours = None
    for r in range(1, 5):
        m = HDR_HOURS.search(n(ws.cell(r, 1).value))
        if m:
            hours = int(m.group(1))
            break
    subs = []
    for r in range(3, ws.max_row + 1):
        v = n(ws.cell(r, 1).value).replace("\n", " ").strip()
        if not v or HDR_HOURS.search(v):
            continue
        m = SUB.match(v)
        if m:
            nm = clean(m.group(1))
            if nm:
                subs.append((nm, int(m.group(2))))
    return big, hours, subs


def build_grade(g):
    path = os.path.join(SRC, f"r6_sansu_nenkei_{g}.xlsx")
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    units, matched, kept = [], 0, 0
    for sn in wb.sheetnames:
        big, hours, subs = parse_sheet(wb[sn])
        if not big or not hours:
            continue
        if subs and sum(h for _, h in subs) == hours and len(subs) >= 2:
            for snm, sh in subs:
                nm = f"{big}「{snm}」"
                units.append({"name": nm, "lessons": [nm] * sh})
            matched += 1
        else:
            units.append({"name": big, "lessons": [big] * hours})
            kept += 1
    return units, matched, kept


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    idx_path = os.path.join(TPL, "index.json")
    idx = json.load(open(idx_path, encoding="utf-8"))
    byid = {t["id"]: t for t in idx["templates"]}
    changed = 0
    for g in range(1, 7):
        r = build_grade(g)
        if not r:
            print(f"  {g}年: 原本なし")
            continue
        units, matched, kept = r
        tot = sum(len(u["lessons"]) for u in units)
        tid = f"nichibun_sansu{g}"
        p = os.path.join(TPL, f"{tid}.json")
        if not os.path.exists(p):
            print(f"  {g}年: テンプレなし")
            continue
        before = json.load(open(p, encoding="utf-8"))
        before_tot = sum(len(u["lessons"]) for u in before["units"])
        big8 = sum(1 for u in units if len(u["lessons"]) >= 8)
        ok = tot == before_tot
        print(
            f"  {g}年: {len(before['units'])}→{len(units)}単元 {tot}時"
            f"（分割{matched}/維持{kept}・8コマ以上{big8}）"
            f"{'✓' if ok else f'✗時数 {before_tot}→{tot}'}"
        )
        if not ok or not args.apply:
            continue
        before["units"] = units
        base = (before.get("note") or "").split("★再構築")[0].rstrip()
        before["note"] = (
            base
            + "★再構築: 各シート内のA列小単元見出し「名称（ページ）N時間」を単元の基準にし、"
            "『大単元「小単元」』の入れ子命名へ。小単元の合計が大単元の配当と一致する場合のみ分割"
            "（不一致は分割せず維持＝推測しない）。原本は sources/teaching-plans に保管。"
        )
        json.dump(before, open(p, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        if tid in byid:
            byid[tid]["units"] = len(units)
            byid[tid]["periods"] = tot
        changed += 1
    if args.apply and changed:
        json.dump(idx, open(idx_path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"\nindex.json 同期・{changed}学年を更新")


if __name__ == "__main__":
    main()
